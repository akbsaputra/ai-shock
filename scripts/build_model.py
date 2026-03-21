from __future__ import annotations

import json
import re
from copy import deepcopy
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

WORKBOOK_DEFAULT = Path(r"c:\Users\akbar\Downloads\AI shock model (11).xlsx")
OUTPUT_PATH = Path("src/model/modelData.json")

SHEET_NAMES = [
    "assumptions",
    "summary",
    "changes",
    "charts",
    "low-low",
    "med-low",
    "high-low",
    "low-high",
    "med-high",
    "high-high",
]

EDITABLE_GROUPS = {
    "Core Drivers": [
        "C10",
        "C14",
        "D14",
        "E14",
        "C15",
        "D15",
        "E15",
        "C16",
        "D16",
        "E16",
        "C17",
        "D17",
        "E17",
        "C23",
        "D23",
        "C24",
        "D24",
        "C25",
        "D25",
    ],
    "Tax & Pass-through": [
        "C5",
        "C6",
        "C7",
        "C8",
        "C27",
        "C30",
        "C31",
        "C32",
        "C35",
    ],
    "Labour Tax Dynamics": [
        "C19",
        "C20",
        "C37",
        "C38",
        "C39",
    ],
}

DERIVED_FIELDS = ["C4", "C11", "C12", "C21", "C29", "C33", "C34"]

PERCENT_RANGE = {"min": 0.0, "max": 1.0, "step": 0.01, "format": "percent"}
SHARE_RANGE = {"min": 0.0, "max": 100.0, "step": 0.1, "format": "points"}
RATE_RANGE = {"min": 0.0, "max": 1.0, "step": 0.005, "format": "percent"}
LENGTH_RANGE = {"min": 0.0, "max": 20.0, "step": 0.5, "format": "number"}

RANGE_MAP = {
    "C5": SHARE_RANGE,
    "C6": SHARE_RANGE,
    "C7": SHARE_RANGE,
    "C8": SHARE_RANGE,
    "C10": PERCENT_RANGE,
    "C14": PERCENT_RANGE,
    "D14": PERCENT_RANGE,
    "E14": PERCENT_RANGE,
    "C15": PERCENT_RANGE,
    "D15": PERCENT_RANGE,
    "E15": PERCENT_RANGE,
    "C16": PERCENT_RANGE,
    "D16": PERCENT_RANGE,
    "E16": PERCENT_RANGE,
    "C17": LENGTH_RANGE,
    "D17": LENGTH_RANGE,
    "E17": LENGTH_RANGE,
    "C19": RATE_RANGE,
    "C20": RATE_RANGE,
    "C23": PERCENT_RANGE,
    "D23": PERCENT_RANGE,
    "C24": PERCENT_RANGE,
    "D24": PERCENT_RANGE,
    "C25": PERCENT_RANGE,
    "D25": PERCENT_RANGE,
    "C27": RATE_RANGE,
    "C30": RATE_RANGE,
    "C31": RATE_RANGE,
    "C32": RATE_RANGE,
    "C35": RATE_RANGE,
    "C37": RATE_RANGE,
    "C38": RATE_RANGE,
    "C39": RATE_RANGE,
}

SCENARIO_KEYS = ["low-low", "med-low", "high-low", "low-high", "med-high", "high-high"]


def normalize_formula(formula: str) -> str:
    return formula


def serialize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def sheet_matrix(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[list[Any]]:
    data: list[list[Any]] = []
    for r in range(1, ws.max_row + 1):
        row: list[Any] = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            if cell.value is None:
                row.append(None)
            elif isinstance(cell.value, str) and cell.data_type == "f":
                if cell.value.startswith("="):
                    row.append(normalize_formula(cell.value))
                else:
                    row.append("=" + normalize_formula(cell.value))
            elif isinstance(cell.value, str) and cell.value.startswith("="):
                row.append(normalize_formula(cell.value))
            else:
                row.append(serialize_value(cell.value))
        data.append(row)
    return data


def build_named_expressions(wb: openpyxl.Workbook) -> dict[str, str]:
    expressions: dict[str, str] = {}
    for name, definition in wb.defined_names.items():
        attr_text = getattr(definition, "attr_text", "")
        if not attr_text:
            continue
        expr = attr_text
        if not expr.startswith("="):
            expr = "=" + expr
        expressions[name] = expr
    return expressions


def parse_cell(address: str) -> tuple[int, int]:
    col, row = coordinate_from_string(address)
    return row, column_index_from_string(col)


def make_control(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    address: str,
    group: str,
) -> dict[str, Any]:
    row_idx, col_idx = parse_cell(address)
    label = ws.cell(row_idx, 1).value
    variable = ws.cell(row_idx, 2).value
    additional_info = ws.cell(row_idx, 6).value

    qualifier = ""
    if col_idx == 4:
        qualifier = " (medium/high or high-market-power)"
    if col_idx == 5:
        qualifier = " (high)"
    if row_idx in {14, 15, 16, 17}:
        qualifier = {3: " (low)", 4: " (medium)", 5: " (high)"}[col_idx]
    if row_idx in {23, 24, 25}:
        qualifier = {3: " (low market power)", 4: " (high market power)"}[col_idx]

    field_range = deepcopy(RANGE_MAP[address])
    default_value = ws.cell(row_idx, col_idx).value

    return {
        "id": f"assumptions-{address.lower()}",
        "sheet": "assumptions",
        "sheetCell": address,
        "group": group,
        "label": f"{label}{qualifier}",
        "variable": variable,
        "additionalInfo": additional_info,
        "editable": True,
        "defaultValue": default_value,
        **field_range,
    }


def build_assumption_meta(
    ws_assumptions: openpyxl.worksheet.worksheet.Worksheet,
) -> dict[str, Any]:
    controls: list[dict[str, Any]] = []
    for group, addresses in EDITABLE_GROUPS.items():
        for address in addresses:
            controls.append(make_control(ws_assumptions, address, group))

    derived: list[dict[str, Any]] = []
    for address in DERIVED_FIELDS:
        row_idx, col_idx = parse_cell(address)
        derived.append(
            {
                "id": f"assumptions-{address.lower()}",
                "sheet": "assumptions",
                "sheetCell": address,
                "label": ws_assumptions.cell(row_idx, 1).value,
                "variable": ws_assumptions.cell(row_idx, 2).value,
                "additionalInfo": ws_assumptions.cell(row_idx, 6).value,
                "editable": False,
            }
        )

    info_map: dict[str, str] = {}
    for row in range(1, ws_assumptions.max_row + 1):
        cell_a = ws_assumptions.cell(row, 1).value
        cell_f = ws_assumptions.cell(row, 6).value
        if isinstance(cell_a, str) and isinstance(cell_f, str) and cell_f.strip():
            info_map[cell_a] = cell_f

    return {
        "controls": controls,
        "derived": derived,
        "groups": list(EDITABLE_GROUPS.keys()),
        "infoMap": info_map,
    }


def values_matrix(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[list[Any]]:
    data: list[list[Any]] = []
    for r in range(1, ws.max_row + 1):
        row: list[Any] = []
        for c in range(1, ws.max_column + 1):
            row.append(serialize_value(ws.cell(r, c).value))
        data.append(row)
    return data


def submatrix(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> list[list[Any]]:
    out: list[list[Any]] = []
    for r in range(min_row, max_row + 1):
        row: list[Any] = []
        for c in range(min_col, max_col + 1):
            row.append(serialize_value(ws.cell(r, c).value))
        out.append(row)
    return out


def build_summary_groups(ws_summary_values: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, Any]]:
    # Rows are inclusive ranges in workbook indexing
    groups = [
        ("Without AI shock", 4, 8),
        ("Scenario labour tax (total)", 10, 15),
        ("Scenario capital income tax (total)", 17, 22),
        ("Scenario consumption tax (total)", 24, 29),
        ("Total tax revenue (after AI shock)", 31, 36),
        ("Scenario benefit spending", 38, 43),
        ("Net fiscal impact due to AI shock", 45, 50),
    ]
    summary_groups: list[dict[str, Any]] = []
    for title, start_row, end_row in groups:
        row_labels: list[str] = []
        for row_idx in range(start_row, end_row + 1):
            label = ws_summary_values.cell(row_idx, 1).value
            row_labels.append(str(label) if label is not None else "")
        summary_groups.append(
            {
                "title": title,
                "startRow": start_row,
                "endRow": end_row,
                "rowLabels": row_labels,
            }
        )
    return summary_groups


def build_scenario_meta(ws_values: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, Any]:
    row_labels = []
    for r in range(3, 31):
        row_labels.append(str(ws_values.cell(r, 1).value or ""))

    section_rows = []
    for r in [3, 8, 13, 17, 22, 25]:
        section_rows.append({"row": r, "label": str(ws_values.cell(r, 1).value or "")})

    return {
        "rowLabelRange": {"start": 3, "end": 30},
        "valueRange": {"startRow": 2, "endRow": 30, "startCol": 2, "endCol": 13},
        "yearHeaderRow": 2,
        "summaryRows": [27, 28, 29, 30],
        "baselineRow": 26,
        "sectionRows": section_rows,
        "rowLabels": row_labels,
    }


def main() -> None:
    workbook_path = WORKBOOK_DEFAULT
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb_formula = openpyxl.load_workbook(workbook_path, data_only=False)
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)

    sheets: dict[str, Any] = {}
    sheets_values: dict[str, Any] = {}

    for name in SHEET_NAMES:
        ws_formula = wb_formula[name]
        ws_values = wb_values[name]
        sheets[name] = {
            "maxRow": ws_formula.max_row,
            "maxCol": ws_formula.max_column,
            "cells": sheet_matrix(ws_formula),
        }
        sheets_values[name] = {
            "maxRow": ws_values.max_row,
            "maxCol": ws_values.max_column,
            "cells": values_matrix(ws_values),
        }

    named_expressions = build_named_expressions(wb_formula)
    ws_assumptions_values = wb_values["assumptions"]
    ws_summary_values = wb_values["summary"]
    ws_changes_values = wb_values["changes"]

    assumptions_meta = build_assumption_meta(ws_assumptions_values)

    scenario_meta = build_scenario_meta(wb_values["low-low"])

    data = {
        "metadata": {
            "title": "AI Shock Fiscal Simulator",
            "generatedAt": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
            "workbookSource": str(workbook_path),
            "workbookLastModified": datetime.fromtimestamp(workbook_path.stat().st_mtime, UTC)
            .isoformat()
            .replace("+00:00", "Z"),
            "includedSheets": SHEET_NAMES,
            "scenarioKeys": SCENARIO_KEYS,
            "excludedSheets": ["diff countries"],
            "units": "Workbook native units",
        },
        "sheets": sheets,
        "sheetsValues": sheets_values,
        "namedExpressions": named_expressions,
        "assumptions": assumptions_meta,
        "summary": {
            "yearHeaderRow": 2,
            "dataRange": {"startRow": 4, "endRow": 50, "startCol": 1, "endCol": 13},
            "groups": build_summary_groups(ws_summary_values),
        },
        "changes": {
            "headerRow": 2,
            "dataRange": {"startRow": 3, "endRow": 8, "startCol": 1, "endCol": 3},
        },
        "scenarios": {
            "keys": SCENARIO_KEYS,
            "sheetByKey": {key: key for key in SCENARIO_KEYS},
            "meta": scenario_meta,
        },
        "baselineSnapshots": {
            "summaryB4L50": submatrix(ws_summary_values, 4, 50, 2, 12),
            "changesB3C8": submatrix(ws_changes_values, 3, 8, 2, 3),
            "scenarioTables": {
                key: submatrix(wb_values[key], 4, 30, 2, 13) for key in SCENARIO_KEYS
            },
        },
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(data, indent=2), encoding="utf-8")
    print(f"Wrote model data to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
